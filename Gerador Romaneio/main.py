from email.policy import default
from tkinter import *
from tkinter import ttk
from typing import Sized
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import os
ferramentas=['','Alicate ','Alicate','Serra','Alicate','Parafusadeira']
unidades=['','1/8"', '1/4"', '1/2"', '1"']
data=datetime.date.today()
dia=datetime.date.day
mes=datetime.date.month
ano=datetime.date.year
itens = []

def Adicionar():
    item = [ferramenta_entry.get(),unidade_combobox.get(),quantidade_spinbox.get()]
    itens.append(item)
    for i in itens:
      print(i)

    


def RemoverItem():
    itens.pop()

def CriaPlanilha():
    # ferramentas.sort()
    wb= Workbook()
    ws = wb.active
    ws.title="Testes"
    row = ws.row_dimensions[1]
    row.font =Font(bold=True)
  
    ws['A1']="Item"
    ws['B1']="QTD"
    ws['C1']="Unidade"
    ws['D1']="Part Nº"
    ws['E1']="Descrição do Material"
   
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    img=Image("Logo.png")
    
 
    ws.add_image(img,'A2')
    for ferramenta in itens:
        ws.append([ferramenta[0],ferramenta[1],ferramenta[2]])
       
        
    
      
    wb.save("Teste.xlsx")



window = Tk()
window.title("Automatização Planilhas Master")

frame = ttk.Frame(window)
frame.pack()

# Frame de informações
user_info_frame =ttk.LabelFrame(frame, text="Gerador romaneio")
user_info_frame.grid(row= 0, column=0, padx=10, pady=20)

# Inputs 
nome_entry = ttk.Combobox(user_info_frame,value=ferramentas)
nome_entry.grid(row=0, column=0)

ferramenta_label = ttk.Label(user_info_frame, text="Ferramenta")
ferramenta_label.grid(row=0, column=0)


ferramenta_entry = ttk.Combobox(user_info_frame,value=ferramentas)
ferramenta_entry.grid(row=1, column=0)

ferramenta_label = ttk.Label(user_info_frame, text="Ferramenta")
ferramenta_label.grid(row=0, column=0)



# itens_adicionados_label=ttk.Label(user_info_frame,text=ferramentas)
# itens_adicionados_label.grid(row=4,column=1)
title_label = ttk.Label(user_info_frame, text="Polegadas")
unidade_combobox = ttk.Combobox(user_info_frame, values=unidades)
title_label.grid(row=0, column=1)
unidade_combobox.grid(row=1, column=1)

quantidade_label = ttk.Label(user_info_frame, text="Quantidade")
quantidade_spinbox = ttk.Spinbox(user_info_frame,from_=1, to="infinity")
quantidade_label.grid(row=0, column=2)
quantidade_spinbox.grid(row=1, column=2)




for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)








# Botões
botaoAdicionarItem = ttk.Button(user_info_frame, text="Adicionar Ferramenta", command= Adicionar)
botaoAdicionarItem.grid(row=4, column=1, sticky="news", padx=20, pady=10,)

botaoRemoverItem = ttk.Button(user_info_frame, text="Remover Ultimo item", command= RemoverItem)
botaoRemoverItem.grid(row=6, column=1, sticky="news", padx=10, pady=10)

botaoCriarPlanilha = ttk.Button(frame, text="Gerar planilha", command= CriaPlanilha)
botaoCriarPlanilha.grid(row=3, column=0, sticky="news", padx=10, pady=10)
 
window.mainloop()
